from __future__ import annotations

from io import BytesIO
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

REPORT_HEADERS = (
    "page",
    "label_index",
    "global_index",
    "order_no",
    "tracking_no",
    "status",
    "message",
    "font_size",
    "overlay_x0",
    "overlay_y0",
    "overlay_x1",
    "overlay_y1",
)


def _build_workbook(report_rows: list[dict[str, Any]]) -> Workbook:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Report"
    sheet.append(REPORT_HEADERS)

    for row in report_rows:
        sheet.append([row.get(header, "") for header in REPORT_HEADERS])

    header_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

    for column_cells in sheet.columns:
        column_letter = get_column_letter(column_cells[0].column)
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        sheet.column_dimensions[column_letter].width = min(max(max_length + 2, 10), 48)

    return workbook


def save_report_excel(report_rows: list[dict[str, Any]], output_path: str | Path) -> Path:
    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = _build_workbook(report_rows)
    workbook.save(path)
    return path


def report_to_excel_bytes(report_rows: list[dict[str, Any]]) -> bytes:
    workbook = _build_workbook(report_rows)
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()

